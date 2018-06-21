import openpyxl, re

pat = re.compile(r'\+?3?8?\(?0[697][35798]\)?\d{3}-?\d{2}-?\d{2}')
def parser(file):
    if file.split('.')[1] in ["xlsx", 'xlx']:
        return list(set(parsxlsx(file)))
    else: raise TypeError

def openx(funk):
    def wraper(file):
        doc = openpyxl.load_workbook(file)
        book = doc.sheetnames
        if len(book) > 1:
            for i in book:
                sh = doc[i]
                return funk(sh,sh.max_column, sh.max_row)
        else:
            sh = doc[doc.sheetnames[0]]
            return funk(sh,sh.max_column,sh.max_row)
    return wraper
@openx
def parsxlsx(sh,col,row):
    return [y for i in [i for i in [re.findall(pat,re.sub(r'[()-]','',re.sub(r' ','',i))) for i in [str(sh.cell(row=y,column=i).value) for i in range(1,col+1) for y in range(1,row+1)]] if i] for y in i]

# def parstxt(file):
#     with open(file, 'r') as f:
#         return [re.sub(r'[\)\(\+\- ]', '' ,y) for i in [re.findall(pat,i) for i in f.read().split("\n")] for y in i]


# a = parser(r"c:\Users\Carl\Desktop\1.xlsx")
# print(a)
# print(len(a))
